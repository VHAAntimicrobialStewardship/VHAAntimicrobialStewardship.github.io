"""
apply-combined-guidance.py

Reads the Combined Guidance column from AntimicrobialStewardshipGuidanceCombined.xlsx
(Mehul/Mahul sheet) and replaces the Text field of matching inpatient (ORZID2) menus
in a station's OMJSON file.

- Keyed by "Source Menu Name (Inpatient)"
- Preserves any leading navigation link already in the menu text
- Formats headings and paragraph breaks for consistent rendering
- Reports how many menus were updated vs unchanged vs unmatched
"""

import json
import os
import re
import openpyxl
from pathlib import Path

ROOT = Path(__file__).parent.parent
XLSX   = ROOT / 'AntimicrobialStewardshipGuidanceCombined.xlsx'
OM_JSON = ROOT / 'stations' / '001-TestStation' / 'TestStationOMJSON.json'

# ── Navigation link that should be preserved at the top of inpatient menus ──
NAV_LINK = '[Help, legend, allergy info, consults, alternative antimicrobials and more \\(Inpt\\)](orzid2-gmenu-abx-general-information-inpatient)'


def load_combined_guidance(xlsx_path):
    """Return dict {inpatient_menu_name: combined_text} from Mehul/Mahul sheet."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = None
    for name in wb.sheetnames:
        if name.strip().lower() in ('mehul', 'mahul'):
            ws = wb[name]
            break
    if ws is None:
        raise RuntimeError(f'Mehul/Mahul sheet not found. Sheets: {wb.sheetnames}')

    headers = {
        str(ws.cell(1, c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value
    }
    comb_col = headers['Combined Guidance']
    menu_col = headers['Source Menu Name (Inpatient)']

    guidance = {}
    for r in range(2, ws.max_row + 1):
        menu = ws.cell(r, menu_col).value
        comb = ws.cell(r, comb_col).value
        if not menu or not comb:
            continue
        menu = str(menu).strip()
        comb = str(comb).strip()
        # Skip placeholder values
        if comb.lower() in ('mehul', 'mahul', ''):
            continue
        guidance[menu] = comb
    return guidance


def format_combined_text(raw: str) -> str:
    """
    Normalise line breaks from the Excel combined text:
    - Headings (##) get a blank line before and after
    - Numbered/bulleted list items stay on individual lines
    - Wrapped prose lines within a paragraph are joined
    - Paragraph separators (blank lines) are kept as \n\n
    """
    # Normalise Windows line endings
    text = raw.replace('\r\n', '\n').replace('\r', '\n')

    # Split into blocks (already separated by \n\n in source)
    blocks = re.split(r'\n{2,}', text)
    result_blocks = []
    for block in blocks:
        block = block.strip()
        if not block:
            continue
        # If the block starts with a heading marker, keep it as-is
        if block.startswith('#'):
            result_blocks.append(block)
            continue
        # If every line of the block starts with a digit/bullet, it's a list — keep \n
        lines = block.split('\n')
        is_list = all(re.match(r'^\d+\.|^[-*]', l.strip()) or l.strip() == '' for l in lines if l.strip())
        if is_list:
            result_blocks.append('\n'.join(l for l in lines if l.strip()))
            continue
        # For mixed blocks (link lines, plain paragraphs), keep as-is
        result_blocks.append(block)

    return '\n\n'.join(result_blocks)


def get_nav_prefix(existing_text: str):
    """Return the leading navigation link if present, else empty string."""
    # The nav link is always the first line if it starts with '['
    first_line = existing_text.split('\n')[0].strip()
    if first_line.startswith('[') and '](orzid2-gmenu-abx-general-information' in first_line:
        return first_line + '\n\n'
    return ''


def main():
    print(f'Loading combined guidance from:\n  {XLSX}\n')
    guidance = load_combined_guidance(XLSX)
    print(f'  → {len(guidance)} menus with combined guidance\n')

    print(f'Loading OMJSON from:\n  {OM_JSON}\n')
    with open(OM_JSON, encoding='utf-8') as f:
        data = json.load(f)

    updated = 0
    unchanged = 0
    no_match = 0

    for menu in data['menus']:
        name = menu['Name']
        if name not in guidance:
            no_match += 1
            continue

        combined_raw = guidance[name]
        formatted = format_combined_text(combined_raw)

        # Preserve any existing nav-link prefix
        nav_prefix = get_nav_prefix(menu.get('Text', ''))
        new_text = nav_prefix + formatted

        if menu['Text'] == new_text:
            unchanged += 1
            continue

        menu['Text'] = new_text
        updated += 1

    print(f'Results:')
    print(f'  Updated  : {updated}')
    print(f'  Unchanged: {unchanged}')
    print(f'  No match : {no_match}')
    print()

    with open(OM_JSON, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f'Saved → {OM_JSON}')


if __name__ == '__main__':
    main()
