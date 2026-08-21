"""Build simplified combined menus and add Combined cross-ref fields."""

import json
import os
import re

import openpyxl


test_path = r"stations\001-TestStation\TestStationOMJSON.json"
combined_xlsx = os.environ["TEMP"] + r"\GuidanceCombined.xlsx"


wb = openpyxl.load_workbook(combined_xlsx, read_only=True, data_only=True)
ws = next(s for s in wb.worksheets if s.title.strip().lower() in ("mehul", "mahul"))
headers = {
    str(ws.cell(1, c).value).strip(): c
    for c in range(1, ws.max_column + 1)
    if ws.cell(1, c).value
}
comb_col = headers["Combined Guidance"]
menu_col = headers["Source Menu Name (Inpatient)"]

guidance = {}
for r in range(2, ws.max_row + 1):
    menu = ws.cell(r, menu_col).value
    comb = ws.cell(r, comb_col).value
    if not menu or not comb:
        continue
    menu = str(menu).strip()
    comb = str(comb).strip()
    if comb.lower() in ("", "mehul", "mahul"):
        continue
    guidance[menu] = comb


def to_combined_name(inpt_name):
    name = inpt_name
    name = re.sub(r"^ORZID2\s+GMENU\s+", "", name)
    name = re.sub(r"^ORZID2\s+", "", name)
    name = re.sub(r"^GMENU\s+", "", name)
    name = re.sub(r"^ABX\s+", "", name)
    name = re.sub(r"\s{2,}", " ", name).strip()
    return name


def format_combined(raw):
    text = raw.replace("\r\n", "\n").replace("\r", "\n")
    blocks = re.split(r"\n{2,}", text)
    result = []
    for block in blocks:
        block = block.strip()
        if block:
            result.append(block)
    return "\n\n".join(result)


with open(test_path, encoding="utf-8") as f:
    data = json.load(f)

existing_names = {m["Name"] for m in data["menus"]}
inpt_by_name = {m["Name"]: m for m in data["menus"]}

new_menus = []
cross_refs_added = 0

for inpt_name, combined_text in guidance.items():
    combined_name = to_combined_name(inpt_name)
    if combined_name in existing_names:
        continue

    combined_menu = {
        "Name": combined_name,
        "Term1": "",
        "Term2": "",
        "Text": format_combined(combined_text),
        "LinkTargets": [],
    }
    new_menus.append(combined_menu)

    if inpt_name in inpt_by_name:
        inpt_by_name[inpt_name]["Combined"] = combined_name
        cross_refs_added += 1

print("New combined menus created:", len(new_menus))
print("Combined cross-refs added:", cross_refs_added)

data["menus"].extend(new_menus)

with open(test_path, "w", encoding="utf-8") as f:
    json.dump(data, f, indent=2, ensure_ascii=False)

print("Total menus now:", len(data["menus"]))
print("Saved.")
